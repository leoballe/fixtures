"""
Servidor Flask para el generador de fixtures con asignación manual.

Rutas:
- /                 → index.html
- /import_teams     → POST CSV
- /generate_parts   → POST config → horarios + lista de partidos
- /generate         → POST (opcional) fixture automático
- /export_pdf       → GET fixture automático a PDF
- /export_manual_pdf→ POST fixture manual a PDF
"""
import re
import os
import csv
from flask import Flask, request, jsonify, send_file

from fixture_generator import (
    read_teams_from_csv,
    generate_fixture,
    export_to_pdf,
    generate_timeslots_list,
    generate_match_list,
    generate_24_team_full_tournament,
    generate_24_team_full_tournament_4x6,
    generate_24_team_full_tournament_seeded,
    generate_23_team_full_tournament,
    generate_23_team_full_tournament_seeded,
    generate_23_team_full_tournament_4x6,
    generate_22_team_full_tournament_4x6,
    generate_22_team_full_tournament,
    generate_21_team_full_tournament_seeded,
    generate_21_team_full_tournament_4x6,
    generate_20_team_full_tournament_seeded,
    generate_20_team_full_tournament_4x6,
    generate_19_team_full_tournament_seeded,
    generate_19_team_full_tournament_4x6,
    generate_18_team_full_tournament_seeded,
    generate_18_team_full_tournament_4x6,
    generate_17_team_full_tournament_seeded,
    generate_17_team_full_tournament_4x6,
    generate_16_team_full_tournament_seeded,
    generate_16_team_full_tournament_4x6,
    generate_15_team_full_tournament_seeded,
    generate_15_team_full_tournament_4x6,
    generate_14_team_full_tournament_seeded,
    generate_14_team_full_tournament_4x6,
    generate_13_team_full_tournament_4x6,
    generate_24_team_full_tournament_6x4_seeded,
    generate_23_team_full_tournament_6x4_seeded,
    generate_22_team_full_tournament_6x4_seeded,
    generate_21_team_full_tournament_6x4_seeded,
    generate_20_team_full_tournament_6x4_seeded,
    generate_19_team_full_tournament_6x4_seeded,
    generate_18_team_full_tournament_6x4_seeded,
    generate_17_team_full_tournament_6x4_seeded,
    Team,
    Match,
)


app = Flask(__name__, static_folder='templates', static_url_path='')

loaded_teams: list[Team] = []
current_schedule: list[Match] = []


# ---------------- PÁGINA PRINCIPAL Y ESTÁTICOS ----------------
@app.route('/')
def index_page():
    index_path = os.path.join(app.static_folder, 'index.html')
    return send_file(index_path)


@app.route('/<path:filename>')
def static_files(filename: str):
    file_path = os.path.join(app.static_folder, filename)
    if os.path.isfile(file_path):
        return send_file(file_path)
    return jsonify({'error': 'Archivo no encontrado'}), 404


# ---------------- CARGA DE EQUIPOS ----------------

@app.route('/import_teams', methods=['POST'])
def import_teams():
    global loaded_teams
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Se requiere un archivo CSV.'}), 400

    tmp_path = os.path.join('/tmp', file.filename or 'equipos.csv')
    file.save(tmp_path)
    try:
        loaded_teams = read_teams_from_csv(tmp_path)
    except Exception as exc:
        return jsonify({'error': f'Error al leer CSV: {exc}'}), 400
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return jsonify({'teams': [t.__dict__ for t in loaded_teams]})

# ---------------- LISTAS POR DEFECTO (CSV en la misma carpeta del proyecto/templates) ----------------

LIST_CSV_FILES = {
    'disciplinas': 'Disciplinas.csv',
    'categorias': 'Categorias.csv',
    'generos': 'genero.csv',
    'modalidades': 'modalidad.csv',
}


def _read_simple_csv_items_from_path(path: str) -> list[str]:
    """
    Lee un CSV simple para completar desplegables.
    Acepta archivos de una columna o varias columnas; toma el primer valor no vacío de cada fila.
    Soporta UTF-8, Latin-1 y CP1252, y separadores coma, punto y coma o tabulación.
    """
    tried_encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    last_exc = None

    for enc in tried_encodings:
        items: list[str] = []
        seen: set[str] = set()
        try:
            with open(path, newline='', encoding=enc) as csvfile:
                sample = csvfile.read(4096)
                csvfile.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
                except Exception:
                    dialect = csv.excel

                reader = csv.reader(csvfile, dialect)
                for row in reader:
                    if not row:
                        continue

                    value = ''
                    for cell in row:
                        cell = (cell or '').strip().strip('"').strip("'")
                        if cell:
                            value = cell
                            break

                    if not value:
                        continue

                    # Si el CSV trae encabezado, no lo cargamos como opción.
                    header_like = value.strip().lower()
                    if not items and header_like in {
                        'disciplina', 'disciplinas',
                        'categoria', 'categoría', 'categorias', 'categorías',
                        'genero', 'género', 'generos', 'géneros',
                        'modalidad', 'modalidades',
                        'nombre', 'item', 'valor', 'value'
                    }:
                        continue

                    key = value.casefold()
                    if key in seen:
                        continue

                    seen.add(key)
                    items.append(value)

            return items
        except Exception as exc:
            last_exc = exc

    raise last_exc or RuntimeError(f'No se pudo leer el archivo {path}')


def _candidate_csv_paths(filename: str) -> list[str]:
    """
    Busca el CSV tanto en la carpeta del index.html (templates) como junto a app.py.
    Esto cubre los dos escenarios habituales del proyecto.
    """
    root_dir = os.path.dirname(os.path.abspath(__file__))
    static_dir = os.path.abspath(app.static_folder)

    candidates = [
        os.path.join(static_dir, filename),
        os.path.join(root_dir, filename),
        os.path.join(root_dir, 'templates', filename),
    ]

    # Fallback case-insensitive para evitar problemas por mayúsculas/minúsculas en Linux/Replit.
    for folder in [static_dir, root_dir, os.path.join(root_dir, 'templates')]:
        try:
            for existing in os.listdir(folder):
                if existing.lower() == filename.lower():
                    candidates.append(os.path.join(folder, existing))
        except Exception:
            pass

    unique: list[str] = []
    for p in candidates:
        if p not in unique:
            unique.append(p)
    return unique


def _find_existing_csv(filename: str) -> str | None:
    for path in _candidate_csv_paths(filename):
        if os.path.isfile(path):
            return path
    return None


@app.route('/competition_lists', methods=['GET'])
def competition_lists():
    """
    Devuelve las listas iniciales para los desplegables del HTML.
    Archivos esperados:
      - Disciplinas.csv
      - Categorias.csv
      - genero.csv
      - modalidad.csv
    """
    data = {
        'disciplinas': [],
        'categorias': [],
        'generos': [],
        'modalidades': [],
        'missing_files': [],
        'errors': {},
    }

    for key, filename in LIST_CSV_FILES.items():
        path = _find_existing_csv(filename)
        if not path:
            data['missing_files'].append(filename)
            continue

        try:
            data[key] = _read_simple_csv_items_from_path(path)
        except Exception as exc:
            data['errors'][filename] = str(exc)
            data[key] = []

    return jsonify(data)

# ---------------- IMPORTAR LISTAS (disciplina, categoría, género, modalidad) ----------------

@app.route('/import_list/<kind>', methods=['POST'])
def import_list(kind: str):
    """
    Importa una lista simple desde un CSV.
    El CSV puede ser:
      - una columna (un ítem por fila)
      - o varias columnas (se toma la primera).

    Se usa para disciplina, categoría, género, modalidad.
    """
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Se requiere un archivo CSV.'}), 400

    tmp_path = os.path.join('/tmp', file.filename or f'{kind}.csv')
    file.save(tmp_path)

    try:
        items = _read_simple_csv_items_from_path(tmp_path)
    except Exception as exc:
        return jsonify({'error': f'Error al leer CSV de {kind}: {exc}'}), 400
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return jsonify({'kind': kind, 'items': items})


# ---------------- HORARIOS + LISTA DE PARTIDOS ----------------

@app.route('/generate_parts', methods=['POST'])
def generate_parts():
    """
    Genera:
      - timeslots: horarios por día/cancha (para las tablas)
      - matches:   lista de partidos sin horario (para la lista lateral)
    """
    global loaded_teams
    if not loaded_teams:
        return jsonify({'error': 'No hay equipos cargados.'}), 400

    data = request.get_json(force=True)
    system = data.get('system', 'rr')
    days = int(data.get('days', 1))
    fields = int(data.get('fields', 1))
    start_time = data.get('start_time', '09:00')
    end_time = data.get('end_time', '18:00')
    match_duration = int(data.get('match_duration', 60))
    home_and_away = bool(data.get('home_and_away', False))

    midday_break = data.get('midday_break')
    if isinstance(midday_break, list) and len(midday_break) == 2:
        midday_break_tuple = (midday_break[0], midday_break[1])
    else:
        midday_break_tuple = None

    # Horarios
    try:
        timeslots = generate_timeslots_list(
            days=days,
            fields=fields,
            start_time=start_time,
            end_time=end_time,
            match_duration=match_duration,
            midday_break=midday_break_tuple,
        )
    except Exception as exc:
        return jsonify({'error': f'Error generando horarios: {exc}'}), 400

    # Partidos
    if len(loaded_teams) == 24 and system == '8x3_sembrado':
        # Usa el fixture completo '8x3 Sembrado': fase de zonas + fase final sembrada
        try:
            matches = generate_24_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 24 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 24 and system == '6x4_sembrado':
        # 24 equipos: fixture completo 6x4 Sembrado según PDF
        try:
            matches = generate_24_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 24 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 23 and system == '6x4_sembrado':
        try:
            matches = generate_23_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 23 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 22 and system == '6x4_sembrado':
        try:
            matches = generate_22_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 22 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 21 and system == '6x4_sembrado':
        try:
            matches = generate_21_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 21 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 20 and system == '6x4_sembrado':
        try:
            matches = generate_20_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 20 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 19 and system == '6x4_sembrado':
        try:
            matches = generate_19_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 19 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 18 and system == '6x4_sembrado':
        try:
            matches = generate_18_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 18 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 17 and system == '6x4_sembrado':
        try:
            matches = generate_17_team_full_tournament_6x4_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 17 equipos (6x4 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 24 and system == '8x3':
        # Usa el fixture completo: fase de zonas + fase final
        try:
            matches = generate_24_team_full_tournament(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 24 equipos: {exc}'}), 400

    elif len(loaded_teams) == 24 and system == '4x6':
        # 24 equipos 4x6: fixture completo según PDF (regular + final)
        try:
            matches = generate_24_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 24 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 23 and system == '8x3_sembrado':
        # 23 equipos: fixture completo '8x3 Sembrado' (fase de zonas + fase final sembrada)
        try:
            matches = generate_23_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 23 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 23 and system == '8x3':
        # 23 equipos: fixture completo (8x3) según PDF base
        try:
            matches = generate_23_team_full_tournament(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 23 equipos: {exc}'}), 400

    elif len(loaded_teams) == 23 and system == '4x6':
        # 23 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_23_EQ.pdf
        try:
            matches = generate_23_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 23 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 22 and system == '4x6':
        # 22 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_22_EQ.pdf
        try:
            matches = generate_22_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 22 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 21 and system == '4x6':
        # 21 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_21_EQ.pdf
        try:
            matches = generate_21_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 21 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 20 and system == '4x6':
        # 20 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_20_EQ.pdf
        try:
            matches = generate_20_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 20 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 19 and system == '4x6':
        # 19 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_19_EQ.pdf
        try:
            matches = generate_19_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 19 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 18 and system == '4x6':
        # 18 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_18_EQ.pdf
        try:
            matches = generate_18_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 18 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 17 and system == '4x6':
        # 17 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_17_EQ.pdf
        try:
            matches = generate_17_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 17 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 18 and system == '4x6':
        # 18 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_18_EQ.pdf
        try:
            matches = generate_18_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 18 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 16 and system == '4x6':
        # 16 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_16_EQ.pdf
        try:
            matches = generate_16_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 16 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 15 and system == '4x6':
        # 15 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_15_EQ.pdf
        try:
            matches = generate_15_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 15 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 14 and system == '4x6':
        # 14 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_14_EQ.pdf
        try:
            matches = generate_14_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 14 equipos (4x6): {exc}'}), 400

    elif len(loaded_teams) == 13 and system == '4x6':
        # 13 equipos: fixture completo (4x6) según PDF 4X6_muestra_gpt_13_EQ.pdf
        try:
            matches = generate_13_team_full_tournament_4x6(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 13 equipos (4x6): {exc}'}), 400
    
    elif len(loaded_teams) == 20 and system == '8x3_sembrado':
        # 20 equipos: fixture completo '8x3 Sembrado' (6 zonas de 3 + 1 zona de 2 con ida/vuelta + fase final)
        try:
            matches = generate_20_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 20 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 19 and system == '8x3_sembrado':
        try:
            matches = generate_19_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 19 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 18 and system == '8x3_sembrado':
        try:
            matches = generate_18_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 18 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 17 and system == '8x3_sembrado':
        try:
            matches = generate_17_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 17 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 16 and system == '8x3_sembrado':
        try:
            matches = generate_16_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 16 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 15 and system == '8x3_sembrado':
        try:
            matches = generate_15_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 15 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 14 and system == '8x3_sembrado':
        try:
            matches = generate_14_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 15 equipos (8x3 Sembrado): {exc}'}), 400

    elif len(loaded_teams) == 21 and system == '8x3_sembrado':
        # 21 equipos: fixture completo '8x3 Sembrado' (7 zonas de 3 + fase final) según cronograma
        try:
            matches = generate_21_team_full_tournament_seeded(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 21 equipos (8x3 Sembrado): {exc}'}), 400
            

    elif len(loaded_teams) == 22 and system in ('8x3', '8x3_sembrado'):
        try:
            matches = generate_22_team_full_tournament(loaded_teams)
        except Exception as exc:
            return jsonify({'error': f'Error generando fixture 22 equipos: {exc}'}), 400
    else:
        try:
            matches = generate_match_list(
                loaded_teams,
                system=system,
                home_and_away=home_and_away,
            )
        except Exception as exc:
            return jsonify({'error': f'Error generando partidos: {exc}'}), 400
    # ✅ Competencia en 4 días: mapear todo lo que estaba en Día 5 a Día 4
    # (Regular queda en días 1-2, inicio fase final día 3, definiciones día 4)
    if days == 4 and isinstance(matches, list):
        for m in matches:
            try:
                if int(m.get('day', 0)) == 5:
                    m['day'] = 4
            except Exception:
                # Si algún match no tiene day válido, lo dejamos como está
                pass

    return jsonify({'timeslots': timeslots, 'matches': matches})


# ---------------- FIXTURE AUTOMÁTICO (OPCIONAL) ----------------

@app.route('/generate', methods=['POST'])
def generate():
    global loaded_teams, current_schedule
    if not loaded_teams:
        return jsonify({'error': 'No hay equipos cargados.'}), 400

    data = request.get_json(force=True)
    system = data.get('system', 'rr')
    days = int(data.get('days', 1))
    fields = int(data.get('fields', 1))
    start_time = data.get('start_time', '09:00')
    end_time = data.get('end_time', '18:00')
    match_duration = int(data.get('match_duration', 60))
    rest = int(data.get('rest', match_duration))
    home_and_away = bool(data.get('home_and_away', False))

    midday_break = data.get('midday_break')
    if isinstance(midday_break, list) and len(midday_break) == 2:
        midday_break_tuple = (midday_break[0], midday_break[1])
    else:
        midday_break_tuple = None

    max_per_day = data.get('max_matches_per_day')
    max_per_day_int = int(max_per_day) if max_per_day is not None else None

    try:
        current_schedule = generate_fixture(
            teams=loaded_teams,
            system=system,
            days=days,
            fields=fields,
            start_time=start_time,
            end_time=end_time,
            match_duration=match_duration,
            rest=rest,
            midday_break=midday_break_tuple,
            home_and_away=home_and_away,
            max_matches_per_day=max_per_day_int,
        )
    except Exception as exc:
        return jsonify({'error': f'Error generando fixture: {exc}'}), 400

    return jsonify({'schedule': [m.__dict__ for m in current_schedule]})


# ---------------- EXPORTACIÓN PDF AUTOMÁTICA ----------------

@app.route('/export_pdf', methods=['GET'])
def export_pdf_route():
    global current_schedule
    if not current_schedule:
        return jsonify({'error': 'No hay un fixture generado.'}), 400

    filename = request.args.get('filename', 'fixture.pdf')
    output_path = os.path.join('/tmp', filename)
    try:
        export_to_pdf(current_schedule, output_path, title='Fixture generado automáticamente')
    except Exception as exc:
        return jsonify({'error': f'Error al exportar PDF: {exc}'}), 500

    return send_file(output_path, as_attachment=True, download_name=filename)


# ---------------- EXPORTACIÓN PDF MANUAL (DRAG & DROP) ----------------

@app.route('/export_manual_pdf', methods=['POST'])
def export_manual_pdf():
    """
    Recibe:
      {
        "schedule": [...],
        "meta": {"disciplina": "...", "categoria": "...", "genero": "...", "modalidad": "..."}
      }
    y genera un PDF A4 vertical con 2 tablas por fila.
    """
    data = request.get_json(force=True)
    schedule_data = data.get('schedule')
    if not schedule_data or not isinstance(schedule_data, list):
        return jsonify({"error": "Se requiere una lista 'schedule' con partidos."}), 400

    meta = data.get('meta') or {}
    disciplina = (meta.get('disciplina') or '').strip()
    categoria  = (meta.get('categoria')  or '').strip()
    genero     = (meta.get('genero')     or '').strip()
    modalidad  = (meta.get('modalidad')  or '').strip()

    try:
        from fpdf import FPDF
    except ImportError:
        return jsonify({"error": "La biblioteca fpdf no está instalada en el servidor."}), 500

        # Agrupar partidos por día y ordenarlos por Nº de partido (como en el PDF).
    # Los BYE se insertan en el día correspondiente y NO llevan hora ni cancha.
    from collections import defaultdict
    import re
    def _pdf_safe(s):
        """
        FPDF clásico suele fallar con algunos Unicode (ej: '…').
        Convertimos a latin-1 y reemplazamos lo que no entra.
        """
        if s is None:
            return ""
        if not isinstance(s, str):
            s = str(s)

        # reemplazos típicos que rompen en FPDF
        s = (s
             .replace("…", "...")      # ellipsis
             .replace("\u2013", "-")   # en dash
             .replace("\u2014", "-")   # em dash
             .replace("\u2212", "-")   # minus
        )

        # latin-1 (fpdf clásico) + reemplazo seguro
        return s.encode("latin-1", "replace").decode("latin-1")
    def _norm(x):
            return (str(x).strip() if x is not None else "")

    def _split_compound_home(home: str, away: str, zone: str):
            h = _norm(home)
            a = _norm(away)
            z = _norm(zone)

            if h and (not a) and (" vs " in h.lower()) and ("zona" in h.lower()):
                mm = re.match(r"^\s*(.*?)\s+vs\s+(.*?)\s*[-–]\s*Zona\s*(.*?)\s*$", h, flags=re.IGNORECASE)
                if mm:
                    h2 = (mm.group(1) or "").strip()
                    a2 = (mm.group(2) or "").strip()
                    z2 = (mm.group(3) or "").strip()
                    if (not z) and z2:
                        z = z2
                    h = h2
                    a = a2

            return h, a, z

    def _strip_zone_prefix(text: str, zone_text: str) -> str:
            t = _norm(text)
            zt = _norm(zone_text)
            if not t or not zt:
                return t
            if t.startswith(zt):
                rest = t[len(zt):].strip()
                rest = re.sub(r"^[\s\.\u2026\-–:]+", "", rest).strip()
                return rest
            return t

    def _normalize_position_row(zone: str, home: str, away: str):
            z = _norm(zone)
            h = _norm(home)
            a = _norm(away)

            if (not z) and (h or a):
                candidate = h or a
                mm = re.match(r"^\s*(\d+º)\s*[\s\.\u2026\-–:]+(.*?)\s*$", candidate)
                if mm:
                    z = (mm.group(1) or "").strip()
                    remainder = (mm.group(2) or "").strip()
                    if h:
                        h = remainder
                    else:
                        a = remainder

            if z and ("º" in z):
                h = _strip_zone_prefix(h, z)
                a = _strip_zone_prefix(a, z)
                if (not h) and a and a.upper() != "BYE":
                    h = a
                    a = ""

            return z, h, a

    by_day = defaultdict(list)
    def _strip_position_prefix(text: str, position: str) -> str:
        t = (text or "").strip()
        p = (position or "").strip()
        if not t or not p:
            return t
        m = re.match(rf"^\s*{re.escape(p)}\s*[\s\.\u2026\-–:]+(.*)$", t)
        if m:
            return (m.group(1) or "").strip()
        return t

    def _normalize_position_row(zone: str, home: str, away: str):
        z = (zone or "").strip()
        h = (home or "").strip()
        a = (away or "").strip()

        # Si zone vacío pero home/away es "23º ... PP41" => extraer zone
        if not z:
            candidate = h or a
            mm = re.match(r"^\s*(\d+º)\s*[\s\.\u2026\-–:]+(.*)$", candidate)
            if mm:
                z = (mm.group(1) or "").strip()
                remainder = (mm.group(2) or "").strip()
                if h:
                    h = remainder
                else:
                    a = remainder

        if z and ("º" in z):
            h = _strip_position_prefix(h, z)
            a = _strip_position_prefix(a, z)

            # Caso típico BYE/informativo: home vacío y away=PPxx => mover PPxx a home
            if a and (not h or h.upper() == "BYE" or h == z):
                h = a
                a = ""

        return z, h, a


    by_day = defaultdict(list)

    for m in (schedule_data or []):
            try:
                d = int(m.get("day") or 0)
            except Exception:
                d = 0

            mm = dict(m or {})
            home = (mm.get("home", "") or "").strip()
            away = (mm.get("away", "") or "").strip()
            zone = (mm.get("zone", "") or "").strip()

            home, away, zone = _split_compound_home(home, away, zone)
            zone, home, away = _normalize_position_row(zone, home, away)

            mm["home"] = home
            mm["away"] = away
            mm["zone"] = zone

            by_day[d].append(mm)


    # ✅ BYE que viene desde el frontend (meta.bye_matches_by_day)
    # Normalizamos para que en el PDF se vea prolijo:
    # - Si llega "X vs Y - Zona Z" todo junto, lo separamos en columnas.
    # - Si es informativo de puesto final (ej "23º ... PP41"), lo mostramos en UNA sola celda (Local)
    #   y dejamos Zona vacía para que no se repita.
    bye_by_day = (meta.get("bye_matches_by_day") or {})
    for d_key, arr in (bye_by_day or {}).items():
        try:
            d = int(d_key)
        except Exception:
            continue

        # Evitar duplicados por número dentro del mismo día
        existing_nums = {str(x.get("number")) for x in by_day.get(d, []) if x.get("number") is not None}

        for bm in (arr or []):
            num = bm.get("number", None)
            num_key = str(num) if num is not None else None
            if num_key is not None and num_key in existing_nums:
                continue

            home = (bm.get("home", "") or "").strip()
            away = (bm.get("away", "") or "").strip()
            zone = (bm.get("zone", "") or "").strip()

            home, away, zone = _split_compound_home(home, away, zone)
            zone, home, away = _normalize_position_row(zone, home, away)

            by_day[d].append({
                "day": d,
                "number": num,
                "time": "",     # sin horario
                "field": "",    # sin cancha
                "home": home,
                "away": away,
                "zone": zone,
                "_is_bye": True,
            })

            if num_key is not None:
                existing_nums.add(num_key)

    def _num_key(m):
        n = m.get("number", None)
        try:
            return int(n)
        except Exception:
            mm = re.search(r"\d+", str(n))
            return int(mm.group(0)) if mm else 10**9

    for d in list(by_day.keys()):
        by_day[d] = sorted(by_day[d], key=_num_key)

    days_sorted = sorted([d for d in by_day.keys() if d != 0])


    # PDF A4 VERTICAL
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=False)

    # Paleta
    primary_blue = (0, 92, 173)
    header_celeste = (185, 217, 235)
    row_alt = (234, 244, 252)

    page_margin = 10
    top_y = 32  # dejamos lugar para título + meta
    table_gap_x = 6
    table_gap_y = 8

    usable_w = pdf.w - 2 * page_margin
    table_w = (usable_w - table_gap_x) / 2.0

    # Columnas base (se escalan para que entren dentro de table_w)
    base_cols = [
        ('Nº', 8),
        ('Hora', 14),
        ('Cancha', 14),
        ('Zona', 14),
        ('Local', 24),
        ('Visitante', 24),
    ]
    base_total = sum(w for _, w in base_cols)
    scale = table_w / base_total  # <- clave para que entren
    cols = [(t, max(6, round(w * scale, 1))) for t, w in base_cols]
    total_cols_width = sum(w for _, w in cols)

    def add_page_with_title():
        pdf.add_page()
        # Título
        pdf.set_fill_color(*primary_blue)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Fixture - Juegos Nacionales Evita", ln=True, align="C", fill=True)

        # Meta debajo del título
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "", 10)
        lineas = [x for x in [
            f"Disciplina: {disciplina}" if disciplina else "",
            f"Categoría: {categoria}" if categoria else "",
            f"Género: {genero}" if genero else "",
            f"Modalidad: {modalidad}" if modalidad else "",
        ] if x]
        if lineas:
            pdf.ln(2)
            pdf.multi_cell(0, 6, "  |  ".join(lineas), align="C")
        pdf.ln(2)
        pdf.set_font("Arial", "", 7)


    def _show_round(zone: str) -> bool:
        z = (zone or "").strip().upper()
        return bool(re.fullmatch(r"[A-H]", z) or z in ("ZA1", "ZA2", "A1", "A2"))

    def draw_day_table(day: int, rows: list, x: float, y: float, box_h: float):
        header_h = 6.0
        col_h = 5.0

        n_rows = len(rows)
        available = max(2.0, box_h - header_h - col_h)

        # Ajuste dinámico: siempre entra (aunque sea chiquito)
        row_h = available / max(1, n_rows)
        row_h = max(1.8, min(4.2, row_h))
        font_size = 6 if row_h < 2.4 else 7

        pdf.set_xy(x, y)

        # Encabezado Día
        pdf.set_fill_color(*primary_blue)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(total_cols_width, header_h, f"Día {day}", border=1, ln=1, align="C", fill=True)

        # Encabezado columnas
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "", font_size)
        pdf.set_fill_color(*header_celeste)
        pdf.set_x(x)
        for title, width in cols:
            pdf.cell(width, col_h, title, border=1, align="C", fill=True)
        pdf.ln()

        even = False
        for m in rows:
            time = _pdf_safe(m.get("time", "") or "")
            field = _pdf_safe(m.get("field", "") or "")
            home = _pdf_safe(m.get("home", "") or "")
            away = _pdf_safe(m.get("away", "") or "")
            zone = _pdf_safe(str(m.get("zone", "") or ""))

            # ✅ Regla general: “partido informativo” (falta un equipo) => SIEMPRE sin hora/cancha
            if str(home).strip() == "" or str(away).strip() == "":
                time = ""
                field = ""
                # y lo tratamos como BYE para estilo/seguridad
                m["_is_bye"] = True
            # ✅ 23 equipos: el “23º … PP41” es informativo. En PDF SIEMPRE sin hora/cancha.
            if str(zone).strip() == "23º" and str(away).strip() == "PP41" and (str(home).strip() == "" or str(home).strip().upper() == "BYE"):
                time = ""
                field = ""
            # ✅ 21 equipos: el “21º … PP30” es informativo. En PDF SIEMPRE sin hora/cancha.
            if str(zone).strip() == "21º" and str(away).strip() == "PP30" and (str(home).strip() == "" or str(home).strip().upper() == "BYE"):
                time = ""
                field = ""
            # ✅ 15 equipos: el “15º … PP42” es informativo. En PDF SIEMPRE sin hora/cancha.
            if str(zone).strip() == "15º" and str(away).strip() == "PP42" and (str(home).strip() == "" or str(home).strip().upper() == "BYE"):
                time = ""
                field = ""

                        # ✅ BYE: se imprime en su posición por Nº (sin horario/cancha) y lo marcamos visualmente
            is_bye = bool(m.get("_is_bye")) or (str(home).strip() == "" or str(away).strip() == "")

            # ✅ Regla general: cualquier BYE / informativo SIEMPRE se exporta sin hora/cancha
            # (aunque lo hayan colocado manualmente en la grilla)
            if is_bye:
                time = ""
                field = ""

                pdf.set_fill_color(240, 240, 240)
                fill = True
            else:
                fill = even
                if fill:
                    pdf.set_fill_color(*row_alt)

            even = not even


            row = [
                str(m.get("number") or ""),  # Nº (usa el número del PDF; no se calcula dinámicamente)
                time,
                field,
                zone,
                home,
                away,
            ]

            pdf.set_x(x)
            for (_, width), value in zip(cols, row):
                pdf.cell(width, row_h, value, border=1, align="C", fill=fill)
            pdf.ln()


    # === 6 días por página: 2 columnas x 3 filas ===
    match_number = 1  # (no lo usamos ahora, pero lo dejamos por compatibilidad)

    # Medidas fijas
    box_h = (pdf.h - page_margin - top_y - (2 * table_gap_y)) / 3.0

    for page_start in range(0, len(days_sorted), 6):
        page_days = days_sorted[page_start:page_start + 6]

        add_page_with_title()

        for idx_day, day in enumerate(page_days):
            row = idx_day // 2   # 0..2
            col = idx_day % 2    # 0..1

            x = page_margin + col * (table_w + table_gap_x)
            y = top_y + row * (box_h + table_gap_y)

            draw_day_table(day, by_day[day], x, y, box_h)

    output_path = os.path.join('/tmp', 'fixture_manual.pdf')
    pdf.output(output_path)
    return send_file(output_path, as_attachment=True, download_name='fixture_manual.pdf')
@app.route('/export_manual_excel', methods=['POST'])
def export_manual_excel():
    """
    Exporta a Excel (.xlsx) el fixture MANUAL (lo que el usuario ubicó en la grilla),
    con un formato muy similar al PDF:
      - Bloques de 2 días por "fila" (Día 1 al lado de Día 2, luego Día 3 al lado de Día 4, etc.)
      - Separación visual con 2 columnas vacías entre tablas
      - Último día solo (si hay cantidad impar de días) queda en la tabla izquierda
      - Incluye disciplina, categoría, género y modalidad
      - NO incluye columna "Ronda"
      - Partidos BYE/informativos salen SIEMPRE sin Hora y sin Cancha

    Colores:
      - Igual al PDF manual (título y "Día": azul; headers: celeste; filas alternadas; BYE gris)

    Recibe:
      {
        "schedule": [...],
        "meta": {
          "disciplina": "...",
          "categoria": "...",
          "genero": "...",
          "modalidad": "...",
          "bye_matches_by_day": { "1":[...], "2":[...], ... }   # opcional
        }
      }
    """
    try:
        data = request.get_json(force=True) or {}
        schedule_data = data.get('schedule') or []
        meta = data.get('meta') or {}

        bye_matches_by_day = meta.get('bye_matches_by_day') or {}

        def _norm(x):
            return (str(x).strip() if x is not None else '')

        def _safe_int(x, default=None):
            try:
                return int(x)
            except Exception:
                return default

        def _is_bye_match(home: str, away: str) -> bool:
            h = _norm(home)
            a = _norm(away)
            if not h or not a:
                return True
            if h.upper() == 'BYE' or a.upper() == 'BYE':
                return True
            return False
        def _split_compound_home(home: str, away: str, zone: str):
                """
                Si llega todo junto en home como: "PP29 vs BYE - Zona LLAVE C"
                separa en columnas (home, away, zone). Si no matchea, deja igual.
                """
                h = _norm(home)
                a = _norm(away)
                z = _norm(zone)

                if h and (not a) and (" vs " in h.lower()) and ("zona" in h.lower()):
                    mm = re.match(r"^\s*(.*?)\s+vs\s+(.*?)\s*[-–]\s*Zona\s*(.*?)\s*$", h, flags=re.IGNORECASE)
                    if mm:
                        h2 = (mm.group(1) or "").strip()
                        a2 = (mm.group(2) or "").strip()
                        z2 = (mm.group(3) or "").strip()
                        if (not z) and z2:
                            z = z2
                        h = h2
                        a = a2

                return h, a, z

        def _strip_zone_prefix(text: str, zone_text: str) -> str:
                """
                Si text viene como: "23º ... PP41" y zone_text="23º",
                devuelve "PP41" (sin el puesto ni separadores).
                """
                t = _norm(text)
                zt = _norm(zone_text)
                if not t or not zt:
                    return t

                if t.startswith(zt):
                    rest = t[len(zt):].strip()
                    rest = re.sub(r"^[\s\.\u2026\-–:]+", "", rest).strip()
                    return rest

                return t

        def _normalize_position_row(zone: str, home: str, away: str):
                """
                Regla: si zone es tipo "23º", el "23º" NO puede aparecer en home/away.
                Además, si queda un solo texto (PPxx) y el otro está vacío, lo dejamos en Local.
                """
                z = _norm(zone)
                h = _norm(home)
                a = _norm(away)

                # Si zone vacío pero home/away es "23º ... PP41" => extraer zone
                if (not z) and (h or a):
                    candidate = h or a
                    mm = re.match(r"^\s*(\d+º)\s*[\s\.\u2026\-–:]+(.*?)\s*$", candidate)
                    if mm:
                        z = (mm.group(1) or "").strip()
                        remainder = (mm.group(2) or "").strip()
                        if h:
                            h = remainder
                        else:
                            a = remainder

                if z and ("º" in z):
                    h = _strip_zone_prefix(h, z)
                    a = _strip_zone_prefix(a, z)

                    # Si quedó PPxx en Visitante y Local vacío, pasarlo a Local
                    if (not h) and a and a.upper() != "BYE":
                        h = a
                        a = ""

                return z, h, a

        import re

        def _strip_position_prefix(text: str, position: str) -> str:
                """
                Si text viene como "23º ... PP41" y position="23º" => devuelve "PP41"
                """
                t = _norm(text)
                p = _norm(position)
                if not t or not p:
                    return t

                m = re.match(rf"^\s*{re.escape(p)}\s*[\s\.\u2026\-–:]+(.*)$", t)
                if m:
                    return (m.group(1) or "").strip()
                return t

        def _normalize_position_row(zone: str, home: str, away: str):
                """
                Regla: si zone es tipo "23º", ese "23º" NO debe aparecer en home/away.
                En informativos típicos, dejamos:
                  zone = "23º"
                  home = "PPxx"
                  away = ""
                """
                z = _norm(zone)
                h = _norm(home)
                a = _norm(away)

                # Si zone viene vacío pero home/away es "23º ... PP41" => extraer zone
                if not z:
                    candidate = h or a
                    mm = re.match(r"^\s*(\d+º)\s*[\s\.\u2026\-–:]+(.*)$", candidate)
                    if mm:
                        z = (mm.group(1) or "").strip()
                        remainder = (mm.group(2) or "").strip()
                        if h:
                            h = remainder
                        else:
                            a = remainder

                # Si zone es un puesto, limpiar prefijo en home/away
                if z and ("º" in z):
                    h = _strip_position_prefix(h, z)
                    a = _strip_position_prefix(a, z)

                    # Caso típico BYE/informativo: home vacío y away=PPxx => mover PPxx a home
                    if a and (not h or h.upper() == "BYE" or h == z):
                        h = a
                        a = ""

                return z, h, a

        def _strip_zone_prefix(text: str, zone_text: str) -> str:
            """
            Si text viene como: "23º ... PP41" y zone_text="23º",
            devuelve "PP41" (sin el puesto ni separadores).
            """
            t = _norm(text)
            zt = _norm(zone_text)
            if not t or not zt:
                return t

            if t.startswith(zt):
                rest = t[len(zt):].strip()
                # sacar separadores típicos: "…", "...", "-", "–", ":" y espacios/puntos
                rest = re.sub(r"^[\s\.\u2026\-–:]+", "", rest).strip()
                return rest

            return t

        def _normalize_position_row(zone: str, home: str, away: str):
            """
            Regla: si zone es tipo "23º", el "23º" NO puede aparecer en home/away.
            - Zone queda igual.
            - Home/Away se limpian si venían con "23º ... XXX".
            """
            z = _norm(zone)
            h = _norm(home)
            a = _norm(away)

            # Caso: el puesto vino metido en home o away y zone está vacío -> extraerlo
            if (not z) and (h or a):
                candidate = h or a
                mm = re.match(r"^\s*(\d+º)\s*[\.\u2026\-–:]+\s*(.*?)\s*$", candidate)
                if mm:
                    z = mm.group(1).strip()
                    remainder = (mm.group(2) or "").strip()
                    if h:
                        h = remainder
                    else:
                        a = remainder

            # Si zone es un puesto ("23º"), limpiar home/away si lo contienen
            if z and ("º" in z):
                h = _strip_zone_prefix(h, z)
                a = _strip_zone_prefix(a, z)

            return z, h, a

        # -------------------------
        # Agrupar por día + forzar BYE sin hora/cancha
        # -------------------------
        by_day = {}

        # 1) Partidos colocados en la grilla
        for m in schedule_data:
            d = _safe_int(m.get('day'), default=0)
            n = _safe_int(m.get('number'))
            t = _norm(m.get('time'))
            f = _norm(m.get('field'))
            z = _norm(m.get('zone'))
            h = _norm(m.get('home'))
            a = _norm(m.get('away'))

                # Normalizar casos raros: "A vs B - Zona X" y "23º ... PP41"
            h, a, z = _split_compound_home(h, a, z)
            # Si llega todo junto: "PP29 vs BYE - Zona LLAVE C" => separar en columnas
            if h and (not a) and (" vs " in h.lower()) and ("zona" in h.lower()):
                mm = re.match(r"^\s*(.*?)\s+vs\s+(.*?)\s*[-–]\s*Zona\s*(.*?)\s*$", h, flags=re.IGNORECASE)
                if mm:
                    h = (mm.group(1) or "").strip()
                    a = (mm.group(2) or "").strip()
                    z2 = (mm.group(3) or "").strip()
                    if (not z) and z2:
                        z = z2

            z, h, a = _normalize_position_row(z, h, a)

            bye_flag = _is_bye_match(h, a)
            if bye_flag:
                t = ''
                f = ''

            by_day.setdefault(d, []).append({
                "day": d,
                "number": n,
                "time": t,
                "field": f,
                "home": h,
                "away": a,
                "zone": z,
                "_is_bye": bye_flag,
            })

        # 2) BYE/informativos en las áreas BYE por día (si el frontend los manda)
        #    Evitar duplicados por número
        for day_str, arr in (bye_matches_by_day or {}).items():
            d = _safe_int(day_str, default=0)
            if not isinstance(arr, list):
                continue

            existing_nums = {x.get("number") for x in by_day.get(d, []) if x.get("number") is not None}

            for bm in arr:
                n = _safe_int(bm.get("number"))
                if n is not None and n in existing_nums:
                    continue

                # Tomar datos del item BYE
                home = _norm(bm.get("home"))
                away = _norm(bm.get("away"))
                zone = _norm(bm.get("zone"))

                # Normalizar casos raros: "A vs B - Zona X" y "23º ... PP41"
                home, away, zone = _split_compound_home(home, away, zone)
                zone, home, away = _normalize_position_row(zone, home, away)

                by_day.setdefault(d, []).append({
                    "day": d,
                    "number": n,
                    "time": "",     # sin horario
                    "field": "",    # sin cancha
                    "home": home,
                    "away": away,
                    "zone": zone,
                    "_is_bye": True,
                })


                if n is not None:
                    existing_nums.add(n)

        # Si no hay días válidos
        days_sorted = sorted([d for d in by_day.keys() if d and d > 0])
        if not days_sorted:
            return jsonify({"error": "No hay partidos para exportar."}), 400

        # Ordenar por número de partido dentro de cada día (igual que el PDF)
        def _num_key(m):
            n = m.get("number")
            return (n is None, n if n is not None else 10**9)

        for d in days_sorted:
            by_day[d] = sorted(by_day[d], key=_num_key)

        # -------------------------
        # Crear Excel con formato tipo PDF
        # -------------------------
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Fixture"

        # Layout:
        # Tabla izquierda: 6 columnas (A-F)
        # Separación: 2 columnas vacías (G-H)
        # Tabla derecha: 6 columnas (I-N)
        LEFT_START = 1          # A
        LEFT_COLS = 6
        GAP_COLS = 2
        RIGHT_START = LEFT_START + LEFT_COLS + GAP_COLS  # I (9)
        RIGHT_COLS = 6
        TOTAL_COLS = LEFT_COLS + GAP_COLS + RIGHT_COLS   # 14 (A..N)

        # Paleta (igual al PDF manual)
        PRIMARY_BLUE = "FF005CAD"  # (0, 92, 173)
        HEADER_CELESTE = "FFB9D9EB" # (185, 217, 235)
        ROW_ALT = "FFEAF4FC"       # (234, 244, 252)
        BYE_GRAY = "FFF0F0F0"      # (240, 240, 240)

        fill_primary = PatternFill("solid", fgColor=PRIMARY_BLUE)
        fill_header = PatternFill("solid", fgColor=HEADER_CELESTE)
        fill_alt = PatternFill("solid", fgColor=ROW_ALT)
        fill_bye = PatternFill("solid", fgColor=BYE_GRAY)

        # Estilos
        title_font = Font(bold=True, size=14, color="FFFFFFFF")
        meta_font = Font(italic=True, size=10)
        day_font = Font(bold=True, size=12, color="FFFFFFFF")
        header_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def set_cell(r, c, value, font=None, align=None, use_border=False, fill=None):
            cell = ws.cell(row=r, column=c, value=value)
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if use_border:
                cell.border = border
            if fill:
                cell.fill = fill
            return cell

        def merge_and_set(r, c1, c2, value, font=None, align=None, fill=None, use_border=False):
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            # aplicar estilo en todas las celdas del merge para que el color "pinte" todo el ancho
            for c in range(c1, c2 + 1):
                set_cell(r, c, None, font=font, align=align, use_border=use_border, fill=fill)
            # valor solo en la primera celda del merge
            ws.cell(row=r, column=c1, value=value)
            return ws.cell(row=r, column=c1)

        # (1) Título
        row = 1
        merge_and_set(row, 1, TOTAL_COLS, "Fixture - Juegos Nacionales Evita", font=title_font, align=center, fill=fill_primary)
        row += 1

        # (2) Meta
        disciplina = _norm(meta.get("disciplina"))
        categoria = _norm(meta.get("categoria"))
        genero = _norm(meta.get("genero"))
        modalidad = _norm(meta.get("modalidad"))

        meta_line = f"Disciplina: {disciplina} | Categoría: {categoria} | Género: {genero} | Modalidad: {modalidad}"
        merge_and_set(row, 1, TOTAL_COLS, meta_line, font=meta_font, align=center)
        row += 2  # deja una línea en blanco

        # Headers (con Zona después de Cancha)
        headers = ["Nº", "Hora", "Cancha", "Zona", "Local", "Visitante"]

        def write_day_block(day_left, day_right):
            nonlocal row

            # fila: Día X / Día Y (azul)
            merge_and_set(row, LEFT_START, LEFT_START + LEFT_COLS - 1, f"Día {day_left}", font=day_font, align=center, fill=fill_primary, use_border=True)
            if day_right is not None:
                merge_and_set(row, RIGHT_START, RIGHT_START + RIGHT_COLS - 1, f"Día {day_right}", font=day_font, align=center, fill=fill_primary, use_border=True)
            row += 1

            # fila: headers (celeste)
            for i, h in enumerate(headers):
                set_cell(row, LEFT_START + i, h, font=header_font, align=center, use_border=True, fill=fill_header)
            if day_right is not None:
                for i, h in enumerate(headers):
                    set_cell(row, RIGHT_START + i, h, font=header_font, align=center, use_border=True, fill=fill_header)
            row += 1

            left_list = by_day.get(day_left, [])
            right_list = by_day.get(day_right, []) if day_right is not None else []

            max_len = max(len(left_list), len(right_list), 1)

            def write_match(r, start_col, m, is_alt_row: bool):
                home = str(m.get("home") or "").strip()
                away = str(m.get("away") or "").strip()

                # ✅ Regla general: si falta un equipo o hay BYE => tratar como BYE/informativo
                is_bye = bool(m.get("_is_bye")) or (home == "" or away == "") or (home.upper() == "BYE" or away.upper() == "BYE")

                # ✅ BYE / informativo SIEMPRE sin hora ni cancha
                time = "" if is_bye else (m.get("time") or "")
                field = "" if is_bye else (m.get("field") or "")

                row_fill = fill_bye if is_bye else (fill_alt if is_alt_row else None)

                vals = [
                    m.get("number"),
                    time,
                    field,
                    m.get("zone") or "",
                    home,
                    away,
                ]
                for i, v in enumerate(vals):
                    set_cell(r, start_col + i, v, align=center, use_border=True, fill=row_fill)


            for i in range(max_len):
                is_alt = (i % 2 == 1)  # como en el PDF (filas alternadas)

                # izquierda
                if i < len(left_list):
                    write_match(row, LEFT_START, left_list[i], is_alt)
                else:
                    for j in range(LEFT_COLS):
                        set_cell(row, LEFT_START + j, "", align=center, use_border=True)

                # derecha
                if day_right is not None:
                    if i < len(right_list):
                        write_match(row, RIGHT_START, right_list[i], is_alt)
                    else:
                        for j in range(RIGHT_COLS):
                            set_cell(row, RIGHT_START + j, "", align=center, use_border=True)

                row += 1

            row += 1  # línea en blanco entre bloques

        # Escribir de a pares: (1-2), (3-4), (5-6)...
        i = 0
        while i < len(days_sorted):
            d1 = days_sorted[i]
            d2 = days_sorted[i + 1] if (i + 1) < len(days_sorted) else None
            write_day_block(d1, d2)
            i += 2

        # Congelar encabezados
        ws.freeze_panes = ws["A4"]

        # -------------------------
        # Auto-ajustar anchos por contenido (sin tocar las 2 columnas de gap)
        # -------------------------
        # (Evitar que el título/meta expandan todo: ignoramos filas 1-2)
        MIN_W = 4
        MAX_W = 40

        def _cell_len(v):
            if v is None:
                return 0
            s = str(v)
            return len(s)

        start_row_for_width = 4
        for col in range(1, TOTAL_COLS + 1):
            # Gap columnas (G-H) fijas
            if col in (LEFT_START + LEFT_COLS, LEFT_START + LEFT_COLS + 1):
                ws.column_dimensions[get_column_letter(col)].width = 3
                continue

            max_len = 0
            for r in range(start_row_for_width, ws.max_row + 1):
                max_len = max(max_len, _cell_len(ws.cell(row=r, column=col).value))

            # un poquito más para respirar
            w = max(MIN_W, min(MAX_W, max_len + 2))
            ws.column_dimensions[get_column_letter(col)].width = w

        out_path = os.path.join("/tmp", "fixture_manual.xlsx")
        wb.save(out_path)

        try:
            return send_file(out_path, as_attachment=True, download_name="fixture_manual.xlsx")
        except TypeError:
            return send_file(out_path, as_attachment=True, attachment_filename="fixture_manual.xlsx")

    except Exception as exc:
        return jsonify({'error': f'Error exportando a Excel: {exc}'}), 400

if __name__ == '__main__':
    # En Replit: puerto 5000, sin reloader para no duplicar procesos
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
